#!/usr/bin/env python3
"""
Extract MPP Locality Listing from xlsx to structured JSON.

Input : ~/Downloads/Locality Listing and Property Usage with State and Parliament Constituency.xlsx
Output: data/localities.json

Parses the single sheet (525 data rows + 1 totals row) into structured records with
derived ward codes (A-Q letter prefix of Code column) and parsed constituency codes.
"""
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: openpyxl required. pip install openpyxl")

HOME = Path(os.path.expanduser("~"))
XLSX_PATH = HOME / "Downloads" / "Locality Listing and Property Usage with State and Parliament Constituency.xlsx"
PROJECT = Path(__file__).resolve().parent.parent
OUT_PATH = PROJECT / "data" / "localities.json"

# Map xlsx ward-prefix letters back to the councillor-ward codes in councillors.json
WARD_CODE_MAP = {
    "A": "A", "B": "B", "D": "D",
    "F": "FG", "G": "FG",
    "H": "H", "I": "I",
    "J": "JL", "L": "JL",
    "K": "K", "M": "M",
    "N": "NPQ", "P": "NPQ", "Q": "NPQ",
    "X": None,  # Commercial/institutional entities, no ward assignment
}

CONSTITUENCY_RE = re.compile(
    r"(?P<state_code>N\.\d+)\s+(?P<state_name>.+?)\s+under\s+(?P<parl_code>P\.\d+)\s+(?P<parl_name>.+?)\s*$"
)


def parse_constituency(text):
    """Parse 'N.19 Mambong under P.198 Puncak Borneo' into structured parts.
    Compound entries are separated by newlines (e.g. cross-boundary localities)."""
    if not text:
        return None
    text = str(text).strip()
    parts = [p.strip() for p in re.split(r"[\r\n]+", text) if p.strip()]
    parsed = []
    for part in parts:
        m = CONSTITUENCY_RE.match(part)
        if m:
            parsed.append({
                "stateCode": m.group("state_code"),
                "stateName": m.group("state_name").strip(),
                "parliamentCode": m.group("parl_code"),
                "parliamentName": m.group("parl_name").strip(),
            })
        else:
            parsed.append({"raw": part})
    return {
        "raw": text,
        "parsed": parsed,
        "compound": len(parsed) > 1,
    }


def derive_ward_code(code):
    """A001 -> A. F003 -> F. Returns both the raw letter and the councillor ward group."""
    if not code:
        return None, None
    code = str(code).strip().upper()
    m = re.match(r"^([A-Z])", code)
    if not m:
        return None, None
    letter = m.group(1)
    ward_group = WARD_CODE_MAP.get(letter)
    return letter, ward_group


def int_or_zero(v):
    if v is None or v == "":
        return 0
    try:
        return int(v)
    except (ValueError, TypeError):
        try:
            return int(float(v))
        except (ValueError, TypeError):
            return 0


def main():
    if not XLSX_PATH.exists():
        sys.exit(f"ERROR: source xlsx not found at {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    sheet = wb.active
    print(f"  Sheet: '{sheet.title}'  rows={sheet.max_row}  cols={sheet.max_column}")

    items = []
    totals_row = None

    for row_idx in range(2, sheet.max_row + 1):
        no, code, desc, constituency, res, com, ind, exe = (
            sheet.cell(row_idx, c).value for c in range(1, 9)
        )

        if code is None and desc is None and (constituency or "").strip().lower() == "total":
            totals_row = {
                "residential": int_or_zero(res),
                "commercial":  int_or_zero(com),
                "industrial":  int_or_zero(ind),
                "exempted":    int_or_zero(exe),
            }
            continue

        if code is None:
            continue

        letter, ward_group = derive_ward_code(code)
        parsed_constituency = parse_constituency(constituency)

        items.append({
            "no": int_or_zero(no) or None,
            "code": str(code).strip(),
            "name": (str(desc).strip() if desc else ""),
            "letter": letter,
            "wardCode": ward_group,
            "constituency": parsed_constituency,
            "residential": int_or_zero(res),
            "commercial":  int_or_zero(com),
            "industrial":  int_or_zero(ind),
            "exempted":    int_or_zero(exe),
        })

    # Derived totals (sum of all items — authoritative regardless of sheet totals row)
    sum_res = sum(it["residential"] for it in items)
    sum_com = sum(it["commercial"]  for it in items)
    sum_ind = sum(it["industrial"]  for it in items)
    sum_exe = sum(it["exempted"]    for it in items)

    # Breakdowns
    by_ward = {}
    by_state = {}
    by_parliament = {}
    by_letter = {}

    for it in items:
        wc = it["wardCode"] or "X"
        by_ward[wc] = by_ward.get(wc, 0) + 1
        by_letter[it["letter"] or "?"] = by_letter.get(it["letter"] or "?", 0) + 1

        c = it["constituency"]
        if c and c.get("parsed"):
            for p in c["parsed"]:
                sc = p.get("stateCode")
                pc = p.get("parliamentCode")
                if sc:
                    by_state.setdefault(sc, {"name": p.get("stateName", ""), "count": 0})["count"] += 1
                if pc:
                    by_parliament.setdefault(pc, {"name": p.get("parliamentName", ""), "count": 0})["count"] += 1

    payload = {
        "generatedAt": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "source": "MPP Locality Listing and Property Usage with State and Parliament Constituency",
        "sourceFile": XLSX_PATH.name,
        "items": items,
        "totals": {
            "localities": len(items),
            "residential": sum_res,
            "commercial":  sum_com,
            "industrial":  sum_ind,
            "exempted":    sum_exe,
            "stateConstituencies":      len(by_state),
            "parliamentConstituencies": len(by_parliament),
        },
        "sheetTotals": totals_row,  # Raw totals row from xlsx (may differ from sum)
        "breakdowns": {
            "byWard":         dict(sorted(by_ward.items())),
            "byLetter":       dict(sorted(by_letter.items())),
            "byState":        dict(sorted(by_state.items())),
            "byParliament":   dict(sorted(by_parliament.items())),
        },
    }

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)

    print(f"  ✓ Wrote {OUT_PATH}")
    print(f"    - {payload['totals']['localities']} localities")
    print(f"    - {sum_res:,} residential / {sum_com:,} commercial / {sum_ind:,} industrial / {sum_exe:,} exempted")
    print(f"    - {len(by_state)} state constituencies / {len(by_parliament)} parliament constituencies")
    print(f"    - ward breakdown: {json.dumps(payload['breakdowns']['byWard'])}")


if __name__ == "__main__":
    main()
